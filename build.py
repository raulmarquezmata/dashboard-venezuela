# -*- coding: utf-8 -*-
"""
build.py — Regenera el dashboard HTML y el inventario Excel desde news.json.

news.json es la ÚNICA fuente de verdad. Este script nunca la modifica:
solo lee y produce los dos entregables, garantizando que siempre coincidan.

Uso:
    python build.py                 # genera ambos
    python build.py --solo-html
    python build.py --solo-excel

Requiere en la misma carpeta:
    plantilla.html   (el dashboard; su array NOTICIAS se reemplaza completo)
    news.json
"""
import sys
import json
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

import config
import common
from common import escapar_js, clave_fecha, clasificar, MESES_NOMBRE

RAIZ = Path(__file__).parent
PLANTILLA = RAIZ / "plantilla.html"
SALIDA_HTML = RAIZ / "salida" / "venezuela_dashboard.html"
SALIDA_XLSX = RAIZ / "salida" / "inventario_noticias_dashboard.xlsx"


# ------------------------------------------------------------------ HTML
def item_a_js(it: dict) -> str:
    sects = ",".join("{label:'%s'}" % escapar_js(s) for s in it.get('sectors', []))
    srcs = ",".join("{text:'%s',url:'%s'}" % (escapar_js(s['text']), s['url'])
                    for s in it.get('sources', []))
    return (
        "{cat:'%s', date:'%s', es:{title:'%s', preview:'%s', body:'%s'}, "
        "en:{title:'%s'}, sectors:[%s], sources:[%s]}"
    ) % (
        it['cat'], escapar_js(it['date']),
        escapar_js(it['es']['title']), escapar_js(it['es'].get('preview', '')),
        escapar_js(it['es'].get('body', '')),
        escapar_js(it.get('en', {}).get('title', '')),
        sects, srcs,
    )


def construir_html(items: list) -> Path:
    if not PLANTILLA.exists():
        raise SystemExit(f"Falta la plantilla: {PLANTILLA}\n"
                         "Copiá el dashboard actual como 'plantilla.html'.")
    html = PLANTILLA.read_text(encoding='utf-8')

    ini = html.index('const NOTICIAS = [')
    fin = html.index('];', ini) + 2
    cuerpo = ",\n\n".join(item_a_js(x) for x in items)
    html = html[:ini] + "const NOTICIAS = [\n" + cuerpo + "\n];" + html[fin:]

    # Sella la fecha de actualización en el encabezado (ES y EN).
    hoy = datetime.now()
    html = _sellar_fecha(html, hoy)

    SALIDA_HTML.parent.mkdir(parents=True, exist_ok=True)
    SALIDA_HTML.write_text(html, encoding='utf-8')

    # Validación mínima: el array debe tener tantos items como el JSON.
    verif = html[html.index('const NOTICIAS = ['):html.index('];', html.index('const NOTICIAS = ['))]
    n = verif.count("{cat:'")
    if n != len(items):
        raise SystemExit(f"ERROR de integridad: {n} items en el HTML vs {len(items)} en el JSON.")
    print(f"  HTML  -> {SALIDA_HTML}  ({n} noticias)")
    return SALIDA_HTML


def _sellar_fecha(html: str, dt: datetime) -> str:
    """Reemplaza 'Actualizado: X' y 'Updated: X' por la fecha de hoy."""
    import re
    es = f"Actualizado: {common.fecha_es(dt)}"
    en_mes = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][dt.month-1]
    en = f"Updated: {en_mes} {dt.day}, {dt.year}"
    html = re.sub(r'Actualizado: \d{1,2} \w{3} \d{4}', es, html)
    html = re.sub(r'Updated: \w{3} \d{1,2}, \d{4}', en, html)
    return html


# ------------------------------------------------------------------ EXCEL
def construir_excel(items: list) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F = config.FUENTE
    hdr_font = Font(name=F, bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor=config.COLOR_NAVY)
    body = Font(name=F, size=9)
    bold = Font(name=F, bold=True, size=10)
    thin = Side(style='thin', color='BBBBBB')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra = PatternFill('solid', fgColor='F5F5F2')
    total_fill = PatternFill('solid', fgColor=config.COLOR_GREY)

    for it in items:
        it['_cat'], it['_sec'] = clasificar(it)
        it['_k'] = clave_fecha(it['date'])
    items.sort(key=lambda x: x['_k'], reverse=True)

    wb = Workbook()

    # --- Hoja 1: Inventario
    ws = wb.active
    ws.title = "Inventario"
    cols = ["#", "Fecha", "Categoría principal", "Categorías secundarias",
            "Titular", "N° fuentes", "Fuentes / enlaces"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, bd
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    for i, it in enumerate(items, 1):
        ws.append([i, it['date'], it['_cat'], it['_sec'], it['es']['title'],
                   len(it.get('sources', [])),
                   " | ".join(s['url'] for s in it.get('sources', []))])
        r = i + 1
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.border = body, bd
            cell.alignment = Alignment(vertical='top', wrap_text=(c in (3, 4, 5, 7)))
            if i % 2 == 0:
                cell.fill = zebra
    for i, w in enumerate([5, 14, 30, 30, 80, 10, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(items)+1}"

    # --- Hoja 2: Resumen categorías
    ws2 = wb.create_sheet("Resumen categorías")
    ws2.append(["Categoría", "Noticias", "% del total", "Enlaces de fuente"])
    for c in range(1, 5):
        cell = ws2.cell(row=1, column=c)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, bd
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cnt = Counter(x['_cat'] for x in items)
    lnk = Counter()
    for x in items:
        lnk[x['_cat']] += len(x.get('sources', []))
    fila_total = len(cnt) + 2
    r = 2
    for k, v in cnt.most_common():
        ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
        ws2.cell(row=r, column=3, value=f"=B{r}/$B${fila_total}").number_format = '0.0%'
        ws2.cell(row=r, column=4, value=lnk[k])
        for c in range(1, 5):
            ws2.cell(row=r, column=c).font = body
            ws2.cell(row=r, column=c).border = bd
        r += 1
    ws2.cell(row=r, column=1, value="TOTAL").font = bold
    ws2.cell(row=r, column=2, value=f"=SUM(B2:B{r-1})").font = bold
    ws2.cell(row=r, column=4, value=f"=SUM(D2:D{r-1})").font = bold
    for c in range(1, 5):
        ws2.cell(row=r, column=c).fill = total_fill
        ws2.cell(row=r, column=c).border = bd
    for i, w in enumerate([38, 12, 12, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # --- Hoja 3: Resumen por mes
    ws3 = wb.create_sheet("Resumen por mes")
    ws3.append(["Mes", "Noticias", "Con 2+ fuentes", "Fuente única"])
    for c in range(1, 5):
        cell = ws3.cell(row=1, column=c)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, bd
    porm = {}
    for x in items:
        m = x['_k'][1]
        d = porm.setdefault(m, [0, 0, 0])
        d[0] += 1
        if len(x.get('sources', [])) >= 2:
            d[1] += 1
        else:
            d[2] += 1
    r = 2
    for m in sorted(porm, reverse=True):
        etiqueta = (MESES_NOMBRE.get(m, '') + " 2026") if m else "Sin fecha precisa"
        ws3.cell(row=r, column=1, value=etiqueta)
        for j, val in enumerate(porm[m], 2):
            ws3.cell(row=r, column=j, value=val)
        for c in range(1, 5):
            ws3.cell(row=r, column=c).font = body
            ws3.cell(row=r, column=c).border = bd
        r += 1
    ws3.cell(row=r, column=1, value="TOTAL").font = bold
    for c in (2, 3, 4):
        col = get_column_letter(c)
        ws3.cell(row=r, column=c, value=f"=SUM({col}2:{col}{r-1})").font = bold
    for c in range(1, 5):
        ws3.cell(row=r, column=c).fill = total_fill
        ws3.cell(row=r, column=c).border = bd
    for i, w in enumerate([24, 12, 16, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # --- Hoja 4: Notas
    ws4 = wb.create_sheet("Notas metodológicas")
    unica = sum(1 for x in items if len(x.get('sources', [])) == 1)
    total_links = sum(len(x.get('sources', [])) for x in items)
    notas = [
        ["Inventario de noticias — Dashboard Venezuela (Grupo Solfín)", ""],
        ["", ""],
        ["Generado", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Noticias", str(len(items))],
        ["Enlaces de fuente", str(total_links)],
        ["Con 2 o más fuentes", f"{len(items)-unica} ({(len(items)-unica)/max(len(items),1):.0%})"],
        ["Con fuente única", f"{unica} ({unica/max(len(items),1):.0%})"],
        ["Fuente del inventario", "news.json — generado automáticamente por build.py. "
                                  "No editar este archivo a mano: se sobrescribe."],
        ["", ""],
        ["Advertencias", ""],
        ["Fechas", "Los ítems cargados solo a partir del slug del enlace, sin acceso al "
                   "artículo, llevan fecha aproximada. No usar como fecha de registro oficial."],
        ["Clasificación", "Asignada por reglas de palabras clave (config.py, sección RULES) sobre "
                          "titulares y etiquetas sectoriales. Un ítem puede pertenecer a varias "
                          "categorías; se listan hasta dos secundarias."],
        ["Fuente única", "Típicamente notas de trámite, perfiles o coberturas sin réplica "
                         "verificable. Filtrar por 'N° fuentes = 1' antes de citar en un informe."],
    ]
    for row in notas:
        ws4.append(row)
    ws4.cell(row=1, column=1).font = Font(name=F, bold=True, size=12, color=config.COLOR_NAVY)
    ws4.cell(row=10, column=1).font = Font(name=F, bold=True, size=10, color=config.COLOR_NAVY)
    for r in range(2, len(notas) + 1):
        for c in (1, 2):
            cell = ws4.cell(row=r, column=c)
            if not cell.font.bold:
                cell.font = body
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws4.column_dimensions['A'].width = 42
    ws4.column_dimensions['B'].width = 95

    SALIDA_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SALIDA_XLSX)
    print(f"  EXCEL -> {SALIDA_XLSX}  ({len(items)} noticias, {total_links} enlaces)")
    return SALIDA_XLSX


# ------------------------------------------------------------------ MAIN
def main():
    args = sys.argv[1:]
    items = common.cargar_noticias()
    print(f"news.json: {len(items)} noticias, "
          f"{sum(len(x.get('sources', [])) for x in items)} enlaces\n")
    if '--solo-excel' not in args:
        construir_html(items)
    if '--solo-html' not in args:
        construir_excel(items)
    print("\nListo. Ambos entregables provienen del mismo news.json, así que no "
          "pueden desincronizarse.")


if __name__ == '__main__':
    main()
